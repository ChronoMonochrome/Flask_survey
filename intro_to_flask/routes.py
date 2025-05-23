#!/usr/bin/env python
# -*- coding: UTF-8 -*-

import os
import io
import random
import time
import json
import zipfile

from collections import OrderedDict
from intro_to_flask import app
from flask import Flask, render_template, render_template_string, request, jsonify, session, send_file, flash, url_for, redirect
from flask_api import status
from flask_babel import gettext, ngettext, lazy_gettext
from flask_login import LoginManager, login_user, logout_user, login_required, UserMixin, current_user
from flask_mail import Message, Mail
from sqlalchemy import Column, Integer, Numeric, String, DateTime, ForeignKey, Table
from sqlalchemy.sql.expression import func

import openpyxl
from openpyxl.workbook import Workbook

import pandas as pd
import numpy as np

from .models import User, SurveySchool, SurveyParents, SurveyTeachers, Student, StudentCompetition, Teacher, Employee, Municipality, Organization, SettlementType, EduOrganizationLevel, EducationType, EduCategory, AgeRange, TeachingExperienceRange, WorkExperienceSubjectRange, TotalLoadRange, LoadOnSubjectRange, TrainingInstitution, TrainingDate, RetrainingInstitution, RetrainingDate
from .forms import signup_form, signin_form

from werkzeug.utils import secure_filename

mail = Mail()

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = ""

@login_manager.user_loader
def load_user(user_id):
	return FlaskUser(user_id)

class FlaskUser(UserMixin):
	def __init__(self,id):
		self.id = id
		user = User.query.filter_by(id = id).first()
		for col in user.__table__.columns.keys():
			setattr(self, col, getattr(user, col))

@login_manager.unauthorized_handler
def unauthorized_callback():
	return redirect("/")

@app.route("/")
def home():
	#session.clear()
	if "messages" in request.args:
		messages = json.loads(request.args['messages']).values()
		return render_template("home.html", messages=messages)
	if "errors" in request.args:
		errors = json.loads(request.args['errors']).values()
		return render_template("home.html", errors=errors)
		
	if current_user.is_authenticated:
		user = User.query.filter_by(login = session["login"]).first()
		school_count = app.db.session.query(func.count('*')).select_from(SurveySchool).join(User).filter(User.login == user.login).first()[0]
		parents_count = app.db.session.query(func.count('*')).select_from(SurveyParents).join(User).filter(User.login == user.login).first()[0]
		teachers_count = app.db.session.query(func.count('*')).select_from(SurveyTeachers).join(User).filter(User.login == user.login).first()[0]

		kwargs = {"surveys_count": {"school": school_count, "parents": parents_count, "teachers": teachers_count}}
		if user.employee_id:
			survees = [o for o in User.query.filter(User.role == "student").all() if not o.name.startswith(" ") and not o.name.startswith("\n")]
			if "show_all" in request.args:
				kwargs["show_all"] = True
			if "survey_name" in request.args and "show_all" in request.args:
				kwargs["survey_name"] = request.args["survey_name"]
				if kwargs["survey_name"] == "school":
					model = SurveySchool
				elif kwargs["survey_name"] == "parents":
					model = SurveyParents
				elif kwargs["survey_name"] == "teachers":
					model = SurveyTeachers
				else:
					return redirect(url_for(".home", errors = json.dumps({"400": "Ошибка 400, неверное имя survey_name"})))

				count_dict = OrderedDict()
				for survee in survees:
					count = app.db.session.query(func.count('*')).select_from(model).join(User).filter(User.name == survee.name).first()[0]
					count_dict[survee.name] = count
				
				kwargs["surveys_count_per_survee"] = count_dict
				#print(count_dict)

			return render_template("home.html", surveys = ["teachers"], survees = survees, **kwargs)
		elif user.student_id:
			return render_template("home.html", **kwargs)

	return render_template("home.html")

@app.route("/about")
def about():
	return render_template("about.html")

def allowed_file(filename):
	return os.path.splitext(filename)[1] in app.config["ALLOWED_EXTENSIONS"]

def ensure_folder_exists(folder):
	if not os.path.isdir(folder):
		os.makedirs(folder)

@app.route("/signup", methods=["GET", "POST"])
#@login_required
def signup():
	form = signup_form.SignupForm()

	if current_user.is_authenticated:
		return redirect(url_for("profile"))

	if request.method == "POST":
		if form.validate() == False:
			return render_template("signup.html", form = form)
		else:
			role = form.role.data

			new_user = User(form.login.data, form.name.data, form.password.data, role)
			if role == "student":
				new_subj = Student(user = new_user)
				student_competitions = StudentCompetition(student = new_subj)
			elif role == "teacher":
				new_subj = Teacher(user = new_user)
				app.db.session.add(new_subj)
			elif role == "employee":
				new_subj = Employee(user = new_user)
				app.db.session.add(new_subj)

			app.db.session.add(new_user)
			app.db.session.commit()
			session["login"] = new_user.login
			login_user(FlaskUser(new_user.id))
			if role == "student":
				return redirect(url_for("survey"))
			return redirect(url_for(".home"))
	elif request.method == "GET":
		return render_template("signup.html", form = form)
		

@app.route("/register_additional")
def register_additional():
	#if not session["login"] == "admin":
	#	return redirect(url_for(".home"))
	
	log_name = OrderedDict([('apan_sosh101', 'Муниципальное казенное общеобразовательное учреждение "Средняя общеобразовательная школа №4" с .Киевка'), ('apan_sosh102', 'Муниципальное казенное общеобразовательное учреждение "Средняя общеобразовательная школа №7" с. Рагули'), ('apan_sosh103', 'Муниципальное казенное общеобразовательное учреждение "Средняя общеобразовательная школа №8" с. Манычское'), ('apan_sosh104', 'Муниципальное казенное общеобразовательное учреждение "Средняя общеобразовательная школа №9" с.Воздвиженское'), ('min_sosh101', 'МБОУ СОШ № 1 г. Минеральные Воды'), ('min_sosh102', 'МКОУ гимназия № 2 г. Минеральные Воды'), ('min_sosh103', 'МБОУ лицей № 3 г. Минеральные Воды'), ('min_sosh104', 'МКОУ СОШ № 5 г. Минеральные Воды'), ('min_sosh105', 'МКОУ СОШ № 6 г. Минеральные Воды'), ('min_sosh106', 'МБОУ СОШ № 7'), ('min_sosh107', 'г. Минеральные Воды'), ('min_sosh108', 'МБОУ СОШ № 20 г. Минеральные Воды'), ('min_sosh109', 'МБОУ гимназия № 103 г. Минеральные Воды'), ('min_sosh110', 'МБОУ лицей № 104 г. Минеральные Воды'), ('min_sosh111', 'МБОУ СОШ № 111 г. Минеральные Воды'), ('min_sosh112', 'МБОУ СОШ № 7 с. М. Колодцы  Минераловодского района'), ('min_sosh113', 'МБОУ СОШ № 11 п. Новотерский  Минераловодского района'), ('min_sosh114', 'МКОУ ООШ № 12 п. Ленинский  Минераловодского района'), ('min_sosh115', 'МКОУ СОШ № 17 с. Сунжа  Минераловодского района'), ('min_sosh116', 'МКОУ ООШ № 25 п. Бородыновка Минераловодского района'), ('nev_sh101', 'муниципальное бюджетное общеобразовательное учреждение средняя общеобразовательная школа № 7 города Невинномысска (МБОУ СОШ № 7 г. Невинномысска) - Город Невинномысск'), ('nev_sh102', 'муниципальное бюджетное общеобразовательное учреждение средняя общеобразовательная школа № 18 с углубленным изучением отдельных предметов города Невинномысска (МБОУ СОШ № 18 г. Невинномысска) - Город Невинномысск')])
	log_pass = OrderedDict([('apan_sosh101', '347834233'), ('apan_sosh102', '346784612'), ('apan_sosh103', '390543298'), ('apan_sosh104', '375892472'), ('min_sosh101', '506387865'), ('min_sosh102', '030950393'), ('min_sosh103', '249839780'), ('min_sosh104', '139760770'), ('min_sosh105', '519029496'), ('min_sosh106', '565471344'), ('min_sosh107', '926756819'), ('min_sosh108', '694888840'), ('min_sosh109', '427072336'), ('min_sosh110', '122902511'), ('min_sosh111', '743745146'), ('min_sosh112', '058373251'), ('min_sosh113', '215073179'), ('min_sosh114', '022668243'), ('min_sosh115', '007104350'), ('min_sosh116', '339290760'), ('nev_sh101', '645437458'), ('nev_sh102', '620460919')])
	
	logins_all = [o.login for o in User.query.all()]
			 
	for log, password in log_pass.items():
		login = log_name[log]
		if not login in logins_all:
			new_user = User(login = log, name =  log_name[log], password = password, role = "student")
			new_subj = Student(user = new_user)
			app.db.session.add(new_user)
	app.db.session.commit()
	return "success!"


@app.route("/profile")
@login_required
def profile():
	user = User.query.filter_by(login = session["login"]).first()

	if user is None:
		return redirect(url_for("signin"))

	return redirect(url_for("survey_teachers"))
	#return redirect(url_for("student_form", id = [user.student_id]))
	
@app.route("/survey_school", methods=["GET", "POST"])
#@login_required
def survey_school():
	user = User.query.filter_by(login = session["login"]).first()
	if request.method == "POST":
		print(request.form)
		survey = SurveySchool()
		for k, v in request.form.items():
			print("set %s.%s = %s" % (str(survey), k, str(1)))
			setattr(survey, k, 1)
		user.survey_schools.append(survey)
		app.db.session.commit()
		messages = json.dumps({"main": "Благодарим за участие в опросе"})
		return redirect(url_for('.home', messages=messages))

	return render_template("survey_school_template1.html")
	
	
@app.route("/survey_parents", methods=["GET", "POST"])
@login_required
def survey_parents():
	user = User.query.filter_by(login = session["login"]).first()
	if request.method == "POST":
		print(request.form)
		#return str(request.form)
		survey = SurveyParents()
		for k, v in request.form.items():
			print("set %s.%s = %s" % (str(survey), k, str(v)))
			setattr(survey, k, v)
		user.survey_parents.append(survey)
		app.db.session.commit()
		messages = json.dumps({"main": gettext("Your data was saved.")})
		return redirect(url_for('.home', messages=messages))

	return render_template("survey_parents_template.html")
	
@app.route("/survey_teachers", methods=["GET", "POST"])
@login_required
def survey_teachers():
	#return redirect(url_for('.home'))
	user = User.query.filter_by(login = session["login"]).first()
	if not user.role == "student":
		return redirect(url_for('.home'))

	#if user.login != "test_user":
	#	return redirect(url_for('.home'))

	if request.method == "POST":
		print(request.form)
		#return str(request.form)
		survey = SurveyTeachers()
		for k, v in request.form.items():
			print("set %s.%s = %s" % (str(survey), k, str(v)))
			setattr(survey, k, v)
		user.survey_teachers.append(survey)
		app.db.session.commit()
		messages = json.dumps({"main": gettext("Благодарим за участие в опросе")})
		return redirect(url_for('.home', messages=messages))

	return render_template("survey_teachers_template.html")

@app.route("/count_surveys", methods=["GET"])
@login_required
def count_surveys():
	if current_user.role == "student":
		return redirect(url_for('.home'))

	if current_user.login.endswith("_adm"):
		orgs = app.db.session.query(User).filter(User.municipality == current_user.municipality, User.role == "student").all()
	elif current_user.login in ["admin", "operator"]:
		orgs = app.db.session.query(User).filter(User.role == "student").all()
	else:
		return redirect(url_for('.home'))

	buf = [["Муниципальное образование", "Образовательная организация", "Количество анкет"]]

	for org in orgs:
		count = app.db.session.query(func.count('*')).select_from(SurveyTeachers).join(User).filter(User.name == org.name, User.municipality == org.municipality).first()[0]
		buf.append([str(org.municipality), org.name, str(count)])

	s = "\n".join([";".join(row) for row in buf]).encode("cp1251")
	output = io.BytesIO(s)
	output.seek(0)

	return send_file(output, as_attachment=True, attachment_filename='count.csv')

@app.route("/signin", methods=["GET", "POST"])
def signin():
	if current_user.is_authenticated:
		redirect(url_for("profile"))

	form = signin_form.SigninForm()

	if request.method == "POST":
		if not form.validate():
			return render_template("signin.html", form = form)

		session["login"] = form.login.data
		user = User.query.filter_by(login = session["login"]).first()
		login_user(FlaskUser(user.id))
		if user.role == "student":
			return redirect(url_for("survey_teachers"))
		return redirect(url_for('.home'))

	elif request.method == "GET":
		return render_template("signin.html", form = form)

@app.route("/organizations/<column_name>", methods=["GET"])
@login_required
def organizations(column_name):
	filter_keys = Organization.__table__.columns.keys()
	if not column_name in filter_keys:
		return status.HTTP_400_BAD_REQUEST
	
	filter_params = {k: request.args.get(k) for k in filter_keys if request.args.get(k)}
	orgs = OrderedDict([(o.id, getattr(o, column_name)) for o in Organization.query.filter_by(**filter_params).all()])
	return jsonify(orgs)

@app.route("/survey_export", methods=["GET"])
@login_required
def survey_export():
	user_login = ""
	if not current_user.is_authenticated:
		return redirect(url_for(".home", errors = json.dumps({"404": "Ошибка 404, не найдено страницы, соответствующей указанному адресу"})))

	if not current_user.login in ["admin", "operator"]:
		return redirect(url_for(".home", errors = json.dumps({"404": "Ошибка 404, не найдено страницы, соответствующей указанному адресу"})))

	user = User.query.filter_by(login = session["login"]).first()
	if not user.employee_id:
		return ""

	if "user_mo" in request.args:
		user_mo = request.args["user_mo"]

	if "user_login" in request.args:
		user_login = request.args["user_login"]

	#if not "survey_name" in request.args:
	#	return redirect(url_for(".home", errors = json.dumps({"404": "Ошибка 404, не найдено страницы, соответствующей указанному адресу"})))


	survey_name = "teachers"
	if survey_name == "school":
		model = SurveySchool
	elif survey_name == "parents":
		model = SurveyParents
	elif survey_name == "teachers":
		model = SurveyTeachers
	else:
		return redirect(url_for(".home", errors = json.dumps({"400": "Ошибка 400, неверное имя survey_name"})))

	survey_keys = [k for k in model.__table__.columns.keys() if k != "id"]
	columns = [getattr(model, a) for a in survey_keys]

	def proceed_single(login):
		survey_results = app.db.session.query(User.id, User.login, User.name, User.municipality, *columns).join(model).filter(User.login == login).all()
		try:
			teachers_num = int(app.db.session.query(User.teachers_num).filter(User.login == login).first()[0])
		except:
			raise
			raise RuntimeError(login)
		keys = ["user_id", "user_login", "user_name", "user_municipality"] + survey_keys

		wb_data = [keys] + survey_results
		filename = "survey_%s_template.xlsx" %survey_name
		wb = openpyxl.load_workbook(os.path.join(app.config["DATAFRAMES_DIR"], filename))
		ws = wb.worksheets[0]
	
		for i, row in enumerate(wb_data):
			if i == 1:
				row = list(row)
				row.append(teachers_num)
			ws.append(row)

		output = io.BytesIO()
		wb.save(output) # Changed to wb.save
		output.seek(0)  # Reset the stream position to the beginning
		return output
		
	if user_login:
		output = proceed_single(user_login)
		return send_file(output, attachment_filename="survey_%s_template.xlsx" %survey_name, as_attachment=True)

	export_zip = io.BytesIO()
	
	survees = [o for o in User.query.filter(User.municipality == user_mo).all() if not "_adm" in o.login]
	with zipfile.ZipFile(export_zip, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
		for n, o in enumerate(survees):
			wb = proceed_single(o.login)
			zf.writestr('%s - %s.xlsx' % (o.login, o.name), wb.read())
			#if n > 10:
			#	break
	export_zip.seek(0)
	return send_file(export_zip, attachment_filename="%s.zip"%user_mo, as_attachment=True)

@app.route("/survey_export_all", methods=["GET"])
@login_required
def survey_export_all():
	user_login = ""
	if not current_user.is_authenticated:
		return redirect(url_for(".home", errors = json.dumps({"404": "Ошибка 404, не найдено страницы, соответствующей указанному адресу"})))

	if not current_user.login in ["admin", "operator"]:
		return redirect(url_for(".home", errors = json.dumps({"404": "Ошибка 404, не найдено страницы, соответствующей указанному адресу"})))

	user = User.query.filter_by(login = session["login"]).first()
	if not user.employee_id:
		return ""

	if "user_login" in request.args:
		user_login = request.args["user_login"]

	survey_name = "teachers"
	if survey_name == "school":
		model = SurveySchool
	elif survey_name == "parents":
		model = SurveyParents
	elif survey_name == "teachers":
		model = SurveyTeachers
	else:
		return redirect(url_for(".home", errors = json.dumps({"400": "Ошибка 400, неверное имя survey_name"})))

	survey_keys = [k for k in model.__table__.columns.keys() if k != "id"]
	columns = [getattr(model, a) for a in survey_keys]

	def proceed_single():
		survey_results = app.db.session.query(User.id, User.login, User.name, User.municipality, *columns).join(model).all()
		keys = ["user_id", "user_login", "user_name", "user_municipality"] + survey_keys
	
		wb_data = [keys] + survey_results
		filename = "survey_%s_template1.xlsx" %survey_name
		wb = openpyxl.load_workbook(os.path.join(app.config["DATAFRAMES_DIR"], filename))
		ws = wb.worksheets[0]
	
		for row in wb_data:
			ws.append(row)

		output = io.BytesIO()
		wb.save(output) # Changed to wb.save
		output.seek(0)  # Reset the stream position to the beginning
		return output
		
	output = proceed_single()
	return send_file(output, attachment_filename="survey_%s_template.xlsx" %survey_name, as_attachment=True)

@app.route("/survey_export1", methods=["GET"])
@login_required
def survey_export1():
	user_login = ""
	if not current_user.is_authenticated:
		return redirect(url_for(".home", errors = json.dumps({"404": "Ошибка 404, не найдено страницы, соответствующей указанному адресу"})))

	if not current_user.login in ["admin", "operator"]:
		return redirect(url_for(".home", errors = json.dumps({"404": "Ошибка 404, не найдено страницы, соответствующей указанному адресу"})))

	user = User.query.filter_by(login = session["login"]).first()
	if not user.employee_id:
		return ""

	if "user_login" in request.args:
		user_login = request.args["user_login"]

	model = SurveyTeachers

	survey_keys = [k for k in model.__table__.columns.keys() if k != "id"]
	columns = [getattr(model, a) for a in survey_keys]

	def proceed_single(municipality):
		survey_results = app.db.session.query(User.id, User.login, User.name, User.municipality, *columns).join(model).filter(User.municipality == municipality).all()
		try:
			teachers_num = int(app.db.session.query(func.sum(User.teachers_num)).filter(User.municipality == municipality).scalar())
		except:
			raise RuntimeError(municipality)
		#return
		keys = ["user_id", "user_login", "user_name", "user_municipality"] + survey_keys
	
		wb_data = [keys] + survey_results
		filename = "survey_teachers_template1.xlsx"
		wb = openpyxl.load_workbook(os.path.join(app.config["DATAFRAMES_DIR"], filename))
		ws = wb.worksheets[0]
	
		for i, row in enumerate(wb_data):
			if i == 1:
				row = list(row)
				row.append(teachers_num)
			ws.append(row)

		output = io.BytesIO()
		wb.save(output) # Changed to wb.save
		output.seek(0)  # Reset the stream position to the beginning
		return output
		
	#if user_login:
	#	output = proceed_single(user_login)
	#	return send_file(output, attachment_filename="survey_teachers_template.xlsx", as_attachment=True)

	export_zip = io.BytesIO()

	#survees = [o for o in User.query.filter(User.student_id != None).all() if not o.name.startswith(" ") and not o.name.startswith("\n")]
	mos = list(set(i[0] for i in app.db.session.query(User.municipality).all() if i[0] and not i[0] == "Тест"))
	mos.sort()
	#return str(mos)
	with zipfile.ZipFile(export_zip, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
		for n, mo in enumerate(mos):
			wb = proceed_single(mo)
			#if not wb:
			#	continue
			zf.writestr('%s.xlsx' % (mo), wb.read())
			#if n > 1:
			#	break
	export_zip.seek(0)
	return send_file(export_zip, attachment_filename="output.zip", as_attachment=True)

@app.route("/excel_template", methods=["GET"])
#@login_required
def excel_template():
	columns = request.args.getlist('columns') or []
	filename = request.args.get("filename") or "template.xlsx"
	print(columns)
	# Create a Pandas dataframe from the data.
	df = pd.DataFrame()
	for col_name in columns:
		df[col_name] = np.nan

	output = io.BytesIO()

	# Use the BytesIO object as the filehandle.
	writer = pd.ExcelWriter(output, engine='xlsxwriter')

	# Write the data frame to the BytesIO object.
	df.to_excel(writer, sheet_name='Sheet1')

	writer.save()
	writer.close()
	output.seek(0)
	return send_file(output, attachment_filename=filename, as_attachment=True)


@app.route("/signout")
@login_required
def signout():
	logout_user()
	del session["login"]
	return redirect(url_for("home"))
